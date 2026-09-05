#!/usr/bin/env python3
"""Generate docs/LIMITS_PROTECTION_TRACKER.xlsx — living limits workbook."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "LIMITS_PROTECTION_TRACKER.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
COSMETIC_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_FILL = PatternFill("solid", fgColor="F2F2F2")


def style_header(ws, row: int, ncol: int) -> None:
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_rows(ws, start_row: int, rows: list[list], section: str | None = None) -> int:
    r = start_row
    if section:
        ws.cell(r, 1, section)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ws.max_column or 19)
        ws.cell(r, 1).fill = SECTION_FILL
        ws.cell(r, 1).font = Font(bold=True)
        r += 1
    for row in rows:
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)
        r += 1
    return r


ECU_HEADERS = [
    "Parameter",
    "Category",
    "Tune map (.pclx)",
    "PCLink enabled",
    "Limit ON",
    "Limit OFF / hysteresis",
    "Unit",
    "RPM condition",
    "MAP / load condition",
    "Other conditions",
    "ECU action when tripped",
    "Cluster red box (0x3EE)",
    "Cluster gauge / color only",
    "Status",
    "Last updated",
    "Notes",
]

ECU_ROWS = [
    # Knock
    [
        "Knock detection threshold",
        "Knock",
        "st185-furyx-base-v1",
        "Y",
        "High (startup map default — severe knock only)",
        "—",
        "Link knock level / %",
        "Per Link knock table",
        "Per Link knock table",
        "Logging + retard always; cut only if severe",
        "Knock retard; may cascade BOOST/FUEL/IGN cut",
        "KNOCK (byte 0); cuts per strategy",
        "N",
        "Active — loose",
        str(date.today()),
        "Raise cut threshold for v1; tighten after dyno",
    ],
    [
        "Knock cut / fuel cut on knock",
        "Knock",
        "st185-furyx-base-v1",
        "Y",
        "Loose — Link startup defaults",
        "—",
        "Strategy",
        "—",
        "—",
        "Do not chase idle knock cuts — fix VE/ign first",
        "Ign/fuel/boost cut per knock strategy",
        "KNOCK + IGN/FUEL/BOOST CUT as applicable",
        "N",
        "Active — loose",
        str(date.today()),
        "",
    ],
    # Boost
    [
        "Overboost / boost limit",
        "Boost",
        "st185-furyx-base-v1",
        "Y",
        "32",
        "—",
        "PSI gauge",
        "—",
        "—",
        "MAC 3-port closed-loop boost control active",
        "Boost cut / WG duty limit",
        "BOOST CUT (byte 3)",
        "Bar red 25 PSI; BG flash 27 PSI (cosmetic)",
        "Active — loose",
        str(date.today()),
        "Tuned ref: 29 PSI in st185-furyx-street-v2+",
    ],
    # ECT
    [
        "Coolant temp (ECT) limit",
        "Coolant",
        "st185-furyx-base-v1",
        "Y",
        "265",
        "255",
        "°F",
        "—",
        "—",
        "No Link warning tier; hard limit only",
        "Engine off — fuel and/or ign cut",
        "FUEL CUT / IGN CUT (bytes 1–2)",
        "Gauge color hot @ 230°F",
        "Active — loose",
        str(date.today()),
        "Tuned ref: 240°F ON / 235°F OFF",
    ],
    # Oil pressure
    [
        "Oil pressure minimum limit",
        "Oil",
        "st185-furyx-base-v1",
        "Y",
        "5",
        "—",
        "PSI",
        "> 4500",
        "MAP > 120 kPa",
        "No warning tier; no red-box oil label",
        "Fuel and/or ign cut",
        "FUEL CUT / IGN CUT (bytes 1–2)",
        "Arc ring red < 25 PSI @ RPM > 2000",
        "Active — loose",
        str(date.today()),
        "Tuned ref: 10 PSI, RPM > 3500, MAP > 90 kPa after hot-idle log",
    ],
    # Fuel pressure
    [
        "Fuel pressure minimum",
        "Fuel",
        "st185-furyx-base-v1",
        "N",
        "Off",
        "—",
        "kPa",
        "—",
        "—",
        "Enable after fuel tune stable",
        "May assert BOOST/FUEL CUT per strategy",
        "BOOST CUT / FUEL CUT if enabled",
        "N",
        "Planned",
        str(date.today()),
        "Set floor after Chase Bays rail pressure confirmed",
    ],
    # Coolant pressure
    [
        "Coolant pressure limit",
        "Coolant",
        "st185-furyx-base-v1",
        "N",
        "Off",
        "—",
        "kPa",
        "—",
        "—",
        "Logging only for v1",
        "Cut or log per programming",
        "FUEL/IGN CUT if limit added — no dedicated label",
        "N",
        "Planned",
        str(date.today()),
        "Optional head-gasket / overpressure protection",
    ],
    # Oil temp
    [
        "Oil temperature limit",
        "Oil",
        "st185-furyx-base-v1",
        "N",
        "Off",
        "—",
        "°F",
        "—",
        "—",
        "No ECU limit on base map",
        "None",
        "N/A",
        "Left ring 235°F on / 232°F off",
        "Off",
        str(date.today()),
        "Cosmetic only",
    ],
    # Lambda
    [
        "Lambda AFR limit",
        "Lambda",
        "st185-furyx-base-v1",
        "N",
        "Off",
        "—",
        "λ",
        "—",
        "—",
        "Lean → knock path; rich → logs/smoke",
        "None direct",
        "N/A",
        "Widget color <0.70 or >1.05",
        "Off",
        str(date.today()),
        "Never cluster red box",
    ],
    # Fuel level
    [
        "Fuel level limit",
        "Fuel",
        "st185-furyx-base-v1",
        "N",
        "Off",
        "—",
        "%",
        "—",
        "—",
        "Never",
        "None",
        "N/A",
        "Arc <20%; hardwired lamp",
        "Off",
        str(date.today()),
        "",
    ],
    # ETB
    [
        "ETB / pedal deviation warning",
        "Throttle",
        "st185-furyx-base-v1",
        "Y",
        "Link default (early)",
        "—",
        "Deviation %",
        "—",
        "—",
        "Two-step: warning then hard cut",
        "Cluster warning first",
        "THROTTLE ERR (byte 5)",
        "N",
        "Active — Link default",
        str(date.today()),
        "Do not loosen — safety critical",
    ],
    [
        "ETB / pedal deviation hard cut",
        "Throttle",
        "st185-furyx-base-v1",
        "Y",
        "Link default (stricter than warning)",
        "—",
        "Deviation %",
        "—",
        "—",
        "Engine shutdown",
        "Fuel and/or ign cut",
        "THROTTLE ERR then FUEL/IGN CUT",
        "N",
        "Active — Link default",
        str(date.today()),
        "",
    ],
    # Sensor fault
    [
        "Sensor fault / improbable value",
        "Sensor",
        "st185-furyx-base-v1",
        "Y",
        "Link fault defaults",
        "—",
        "Fault code",
        "—",
        "—",
        "ECT/IAT/MAP/oil/etc. out of range or lost signal",
        "May cut or derate per fault",
        "SENSOR ERR (byte 4)",
        "N",
        "Active",
        str(date.today()),
        "Often accompanies another cut",
    ],
    # Rev limit
    [
        "Rev limit (hard)",
        "Rev",
        "st185-furyx-base-v1",
        "Y",
        "8000",
        "—",
        "RPM",
        "—",
        "—",
        "Motorsport ceiling — NOT a cluster warning label",
        "Ign/fuel cut at limit",
        "Not a dedicated REV LIMIT overlay",
        "N",
        "Active",
        str(date.today()),
        "Use GP soft limit 4000 first idle",
    ],
    [
        "GP / soft RPM limit (first start)",
        "Rev",
        "st185-furyx-base-v1",
        "Y",
        "4000",
        "—",
        "RPM",
        "—",
        "—",
        "Operator first-start / until oil stable",
        "Ign/fuel cut",
        "Not cluster warning",
        "N",
        "Active",
        str(date.today()),
        "See OPERATOR_FIRST_START.md",
    ],
    # TC
    [
        "Traction control intervention",
        "TC",
        "st185-furyx-base-v1",
        "Y",
        "Slip tables (conservative)",
        "—",
        "Slip % / angle",
        "4× wheel speed",
        "Driven vs reference",
        "TC may use ign/boost/fuel internally",
        "Orange TC UI only — suppress red cut bytes during TC",
        "Orange overlay (when implemented)",
        "Planned",
        str(date.today()),
        "0x3ED slip map from cluster encoder",
    ],
    # A/C (load shed — not protection overlay)
    [
        "A/C compressor cut (load shed)",
        "A/C",
        "st185-furyx-base-v1",
        "Y",
        "MAP > 17 PSI OR RPM > 5000",
        "MAP < 14 AND RPM < 4500 for ≥3 s",
        "PSI / RPM",
        "—",
        "—",
        "Not a cluster alarm",
        "A/C clutch off",
        "N/A",
        "N",
        "Active",
        str(date.today()),
        "See FEATURES_AC_IDLE_CRUISE_TC.md",
    ],
]

COSMETIC_HEADERS = [
    "Display",
    "Location",
    "ON threshold",
    "OFF / hysteresis",
    "Unit",
    "Source",
    "ECU limit linked?",
    "Notes",
    "Last updated",
]

COSMETIC_ROWS = [
    ["Boost bar fill red", "Right cluster", "25", "—", "PSI", "Local UI", "N", "Cosmetic only", str(date.today())],
    ["Boost BG flash latch", "Right cluster", "27", "25", "PSI", "Local UI", "N", "Release below 25", str(date.today())],
    ["Coolant gauge color hot", "Center/left", "230", "—", "°F", "ECT on 0x3E8", "N", "ECT limit separate", str(date.today())],
    ["Oil pressure arc/ring", "Left cluster", "25", "—", "PSI @ RPM>2000", "0x3E9 + local", "N", "May flash on loose idle", str(date.today())],
    ["Oil temp ring", "Left cluster", "235", "232", "°F", "0x3E8 + local", "N", "", str(date.today())],
    ["Lambda widget color", "Right cluster", "0.70 / 1.05", "—", "λ", "0x3EA + local", "N", "Bands not alarms", str(date.today())],
    ["Fuel level arc", "Left cluster", "20", "—", "%", "0x3EB + local", "N", "Hardwired lamp too", str(date.today())],
]

STREET_HEADERS = [
    "Cap",
    "Value",
    "Unit",
    "Phase",
    "Notes",
    "Last updated",
]

STREET_ROWS = [
    ["Max boost target (table)", "18", "PSI", "Pre-dyno street", "Not ECU cut — target table", str(date.today())],
    ["First outing RPM cap", "4500", "RPM", "First drives", "Operator / GP limit", str(date.today())],
    ["After stable oil temp RPM", "6000", "RPM", "Street", "", str(date.today())],
    ["Pre-dyno RPM max", "7000", "RPM", "Street", "", str(date.today())],
    ["WOT below RPM", "Avoid sustained", "—", "Street", "Below 3500 RPM", str(date.today())],
]

CHANGELOG_HEADERS = ["Date", "Map file", "Parameter changed", "Old value", "New value", "Reason", "Updated by"]

CHANGELOG_ROWS = [
    [str(date.today()), "st185-furyx-base-v1", "(initial workbook)", "—", "See ECU sheet", "Base startup loose limits", "Agent"],
]


def build() -> None:
    wb = Workbook()

    # --- Sheet 1: README ---
    ws0 = wb.active
    ws0.title = "README"
    readme = [
        ["ST185 XtremeX — Limits snapshot (startup map prep only)"],
        [""],
        ["GOAL: Build st185-furyx-base-v1.pclx — safe startup tune, NOT peak performance."],
        ["Once tuning in PCLink, use PCLink for all limits/tables. This workbook is optional"],
        ["reference while applying the first map. You do not need to keep it updated."],
        [""],
        ["Source: config/limits.yaml + protection docs"],
        [""],
        ["Sheets: ECU Protection | Cluster Cosmetic | Street caps | Changelog (optional)"],
        [""],
        ["Red box: KNOCK, IGN CUT, FUEL CUT, BOOST CUT, SENSOR ERR, THROTTLE ERR only"],
    ]
    for i, row in enumerate(readme, start=1):
        ws0.cell(i, 1, row[0])
    ws0.cell(1, 1).font = Font(bold=True, size=14)
    ws0.column_dimensions["A"].width = 90

    # --- Sheet 2: ECU Protection ---
    ws1 = wb.create_sheet("ECU Protection")
    for c, h in enumerate(ECU_HEADERS, start=1):
        ws1.cell(1, c, h)
    style_header(ws1, 1, len(ECU_HEADERS))
    for r, row in enumerate(ECU_ROWS, start=2):
        for c, val in enumerate(row, start=1):
            ws1.cell(r, c, val)
            ws1.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws1.freeze_panes = "A2"
    set_widths(ws1, [28, 12, 22, 10, 14, 12, 8, 14, 16, 22, 24, 22, 22, 14, 12, 36])

    # --- Sheet 3: Cluster Cosmetic ---
    ws2 = wb.create_sheet("Cluster Cosmetic")
    for c, h in enumerate(COSMETIC_HEADERS, start=1):
        ws2.cell(1, c, h)
    style_header(ws2, 1, len(COSMETIC_HEADERS))
    for r, row in enumerate(COSMETIC_ROWS, start=2):
        for c, val in enumerate(row, start=1):
            cell = ws2.cell(r, c, val)
            cell.fill = COSMETIC_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.freeze_panes = "A2"
    set_widths(ws2, [24, 14, 12, 14, 8, 12, 10, 28, 12])

    # --- Sheet 4: Street / Operator ---
    ws3 = wb.create_sheet("Street Operator Caps")
    for c, h in enumerate(STREET_HEADERS, start=1):
        ws3.cell(1, c, h)
    style_header(ws3, 1, len(STREET_HEADERS))
    for r, row in enumerate(STREET_ROWS, start=2):
        for c, val in enumerate(row, start=1):
            ws3.cell(r, c, val)
    ws3.freeze_panes = "A2"
    set_widths(ws3, [28, 14, 8, 14, 32, 12])

    # --- Sheet 5: Changelog ---
    ws4 = wb.create_sheet("Changelog")
    for c, h in enumerate(CHANGELOG_HEADERS, start=1):
        ws4.cell(1, c, h)
    style_header(ws4, 1, len(CHANGELOG_HEADERS))
    for r, row in enumerate(CHANGELOG_ROWS, start=2):
        for c, val in enumerate(row, start=1):
            ws4.cell(r, c, val)
    ws4.freeze_panes = "A2"
    set_widths(ws4, [12, 24, 28, 14, 14, 36, 14])
    # blank rows for user to fill
    for r in range(len(CHANGELOG_ROWS) + 2, len(CHANGELOG_ROWS) + 52):
        for c in range(1, len(CHANGELOG_HEADERS) + 1):
            ws4.cell(r, c, "")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
